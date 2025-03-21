import openpyxl as op
import calendar
import locale
import os
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment


locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')


def get_or_create_task_path(task: str):
    fecha = datetime.now().strftime('%Y%m')
    input_folder = os.path.join('input')
    output_folder = os.path.join('output')
    input_task_path = os.path.join(input_folder, task, fecha)
    output_task_path = os.path.join(output_folder, task, fecha)
    
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    if not os.path.exists(input_task_path):
        os.makedirs(input_task_path)
    if not os.path.exists(output_task_path):
        os.makedirs(output_task_path)
    
    return input_task_path, output_task_path


def get_date(mes_anterior=False) -> tuple[str, str]:
    """Devuelve el mes y año actual o del mes anterior en formato de texto y número."""
    hoy = datetime.now()
    
    if mes_anterior:
        if hoy.month == 1:
            mes = 12
            año = hoy.year - 1
        else:
            mes = hoy.month - 1
            año = hoy.year
    else:
        mes = hoy.month
        año = hoy.year
    
    mes_nombre = calendar.month_abbr[mes].upper()[:3] # ENE
    mes_año = f"{mes_nombre}{str(año)[2:]}" # ENE24
    fecha = f"{año}{str(mes).zfill(2)}" # 202401
    
    return mes_año, fecha


def get_last_date_pagos(folder_path: str) -> str | None:
    file_list = []
    dates_list = []
    
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx') and 'Base Pagos' in file:
            file_list.append(file)
    
    if not file_list:
        return None
    
    for file in file_list:
        fecha_list = str(file).split(' ')[-1].split('.')[:3]
        dates_list.append(datetime.strptime(f'{fecha_list[0]}.{fecha_list[1]}.{fecha_list[2]}', '%d.%m.%Y'))
    
    last_date: datetime = dates_list[0]
    for date in dates_list:
        if date > last_date:
            last_date = date
    
    return last_date.strftime('%d.%m.%Y')


def get_or_create_folder(folder_path: str) -> str:
    """Crea un directorio si no existe y devuelve su ruta."""
    fecha_hoy = datetime.today().strftime('%Y.%m.%d')
    carpeta_hoy = os.path.join(folder_path, fecha_hoy)
    
    if not os.path.exists(carpeta_hoy):
        os.makedirs(carpeta_hoy)
    
    if not os.path.exists(f'{carpeta_hoy}/agencias'):
        os.makedirs(f'{carpeta_hoy}/agencias')
    
    return carpeta_hoy


def get_hour_am_pm() -> str:
    """Devuelve la hora actual + 1, redondeando a la hora siguiente pasado 30 minutos, en formato de 12 horas con AM/PM."""
    hora = datetime.now() + timedelta(hours=1)
    minuto_ajustado = 30 if 0 < hora.minute <= 30 else 0
    hora = hora.replace(minute=minuto_ajustado, second=0, microsecond=0)
    
    if minuto_ajustado == 0:
        hora += timedelta(hours=1)
    
    periodo = 'PM' if hora.hour >= 12 else 'AM'
    
    return f'{hora.strftime('%H:%M')} {periodo}'


def clean_columns(columns_list: list[str]) -> list[str]:
    """Limpia los nombres de las columnas de un DataFrame."""
    return [column.strip().replace('.', '').replace(' ', '_').upper() for column in columns_list]


def start_file(files: str|list[str]) -> None:
    """Abre un archivo o una lista de archivos."""
    if isinstance(files, list):
        for file in files:
            os.startfile(file)
    else:
        os.startfile(files)


def format_excel(file_path: str, validator: str) -> None:
    """Aplica estilos a un archivo de Excel según el validador seleccionado."""
    workbook = op.load_workbook(file_path)
    sheet = workbook.active
    
    # Definir estilos generales
    general_font = Font(name='Calibri', size=11)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    alignment_center = Alignment(horizontal='center', vertical='center')
    
    # Definir estilos específicos
    header_font_white = Font(name='Calibri', size=11, bold=True, color='000000')
    header_fill_blue = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_fill_yellow = PatternFill(start_color='FFD965', end_color='FFD965', fill_type='solid')
    header_fill_green = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')
    header_fill_orange = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
    
    # Aplicar estilos generales a todas las celdas
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = general_font
    
    # Aplicar estilos al encabezado (fila 1)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = alignment_center
    
    # Aplicar estilos específicos según el validador
    if validator == 'mono' or validator == 'multi' or validator == 'env' or validator == 'noenv':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 12):  # Columnas J-K
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(12, 16):  # Columnas L-O
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(16, 17):  # Columna P
            sheet.cell(row=1, column=col).fill = header_fill_orange
            sheet.cell(row=1, column=col).font = header_font_white
    elif validator == 'multi_agencias':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 13):  # Columnas J-L
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(13, 16):  # Columnas M-O
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
    
    elif validator == 'react_agencias':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 13):  # Columnas J-L
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(13, 16):  # Columnas M-O
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
    
    elif validator == 'reactiva':
        for col in range(1, 10):  # Columnas A-I
            sheet.cell(row=1, column=col).fill = header_fill_blue
        for col in range(10, 13):  # Columnas J-L
            sheet.cell(row=1, column=col).fill = header_fill_green
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(13, 15):  # Columnas M-N
            sheet.cell(row=1, column=col).fill = header_fill_yellow
            sheet.cell(row=1, column=col).font = header_font_white
        for col in range(15, 16):  # Columna O
            sheet.cell(row=1, column=col).fill = header_fill_orange
            sheet.cell(row=1, column=col).font = header_font_white
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(file_path)

